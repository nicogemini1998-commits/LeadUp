import re
import csv
from openpyxl import load_workbook
from io import BytesIO

# Mapeo completo de columnas reales del CSV (nombre normalizado → campo interno)
COLUMN_MAPPING = {
    # Contacto
    'first name':                  'first_name',
    'last name':                   'last_name',
    'contact name':                'contact_name',
    'full name':                   'contact_name',
    'nombre':                      'contact_name',
    'work email':                  'contact_email',
    'email':                       'contact_email',
    'job title':                   'contact_title',
    'title':                       'contact_title',
    'cargo':                       'contact_title',
    'departments':                 'departments',
    'linkedin url':                'linkedin_url',
    'linkedin profile':            'linkedin_url',
    'phone 1':                     'contact_phone',
    'phone1':                      'contact_phone',
    'mobile':                      'contact_phone',
    'direct phone':                'contact_phone',
    'phone 2':                     'contact_phone2',
    'phone2':                      'contact_phone2',
    'telefono':                    'contact_phone',
    # Empresa
    'company name':                'company_name',
    'empresa':                     'company_name',
    'company domain':              'company_domain',
    'company website':             'website',
    'website':                     'website',
    'company city':                'city',
    'city':                        'city',
    'ciudad':                      'city',
    'country':                     'country',
    'pais':                        'country',
    'company main industry':       'industry',
    'industry':                    'industry',
    'industria':                   'industry',
    'company sub industry':        'sub_industry',
    'sub industry':                'sub_industry',
    'company intent topics':       'intent_topics',
    'company number of employees': 'employees',
    'company year founded':        'year_founded',
    'company linkedin url':        'company_linkedin',
    'phone':                       'company_phone',
}

def normalize_column(col_name: str) -> str:
    normalized = col_name.strip().lower()
    normalized = re.sub(r'\s+', ' ', normalized)
    return normalized

def detect_columns(headers: list[str]) -> dict[str, str]:
    detected = {}
    for header in headers:
        norm = normalize_column(header)
        if norm in COLUMN_MAPPING:
            detected[header] = COLUMN_MAPPING[norm]
    return detected

def _combine_name(row_dict: dict) -> dict:
    """Si hay first_name + last_name, combinarlos en contact_name."""
    first = row_dict.pop('first_name', '') or ''
    last = row_dict.pop('last_name', '') or ''
    if first or last:
        combined = f"{first} {last}".strip()
        if combined and not row_dict.get('contact_name'):
            row_dict['contact_name'] = combined
    return row_dict

def parse_csv(file_bytes: bytes) -> dict:
    try:
        try:
            text = file_bytes.decode('utf-8-sig')
        except UnicodeDecodeError:
            text = file_bytes.decode('latin-1')

        lines = text.strip().split('\n')
        reader = csv.DictReader(lines)

        if not reader.fieldnames:
            return {'error': 'CSV vacío o sin cabeceras'}

        headers = list(reader.fieldnames)
        column_mapping = detect_columns(headers)

        data_rows = []
        for row in reader:
            if not any(v.strip() for v in row.values() if v):
                continue
            row_dict = {}
            for header, value in row.items():
                if value and value.strip():
                    internal_field = column_mapping.get(header)
                    if internal_field:
                        row_dict[internal_field] = str(value).strip()
            row_dict = _combine_name(row_dict)
            if row_dict.get('company_name') or row_dict.get('contact_name'):
                data_rows.append(row_dict)

        return {
            'error': None,
            'headers': headers,
            'columns_found': column_mapping,
            'rows': data_rows,
            'total': len(data_rows),
        }
    except Exception as e:
        return {'error': f'Error parsing CSV: {str(e)}'}

def parse_excel(file_bytes: bytes) -> dict:
    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows or not rows[0]:
            return {'error': 'Excel vacío'}

        headers = [str(h).strip() if h else '' for h in rows[0]]
        column_mapping = detect_columns(headers)

        data_rows = []
        for row in rows[1:]:
            if not any(row):
                continue
            row_dict = {}
            for idx, header in enumerate(headers):
                value = row[idx] if idx < len(row) else None
                internal_field = column_mapping.get(header)
                if internal_field and value is not None and str(value).strip():
                    row_dict[internal_field] = str(value).strip()
            row_dict = _combine_name(row_dict)
            if row_dict.get('company_name') or row_dict.get('contact_name'):
                data_rows.append(row_dict)

        return {
            'error': None,
            'headers': headers,
            'columns_found': column_mapping,
            'rows': data_rows,
            'total': len(data_rows),
        }
    except Exception as e:
        return {'error': f'Error parsing Excel: {str(e)}'}
