from __future__ import annotations
import logging
import time
from datetime import date
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from pydantic import BaseModel

from auth import get_current_user, require_admin
from database import get_conn
from services.scheduler import trigger_manual_assignment, assign_leads_for_user
from services.enrichment import enrich_company
from services.langfuse_obs import track_anthropic_call

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/admin", tags=["admin"])

# Cache para rejection analysis: key -> (timestamp, result)
_rejection_cache: dict = {}
_REJECTION_CACHE_TTL = 3600  # 1 hora


class LeadSearchToggle(BaseModel):
    user_id: int
    enabled: bool


class AssignNowRequest(BaseModel):
    user_id: Optional[int] = None
    count: int = 20
    industry: Optional[str] = None


@router.post("/assign-now")
async def assign_now(
    body: AssignNowRequest,
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(require_admin),
):
    """Trigger immediate lead assignment for all users or a specific user."""
    if body.user_id:
        async with get_conn() as conn:
            cursor = await conn.execute(
                "SELECT id, name FROM lu_users WHERE id = ?", (body.user_id,)
            )
            user = await cursor.fetchone()
        if not user:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")

        assigned = await assign_leads_for_user(body.user_id, str(date.today()), body.count, body.industry)
        return {"message": f"Asignados {assigned} leads a {user['name']}", "user_id": body.user_id, "assigned": assigned}

    background_tasks.add_task(trigger_manual_assignment)
    return {"message": "Asignación masiva iniciada para todos los usuarios activos"}

class UserSettingsRequest(BaseModel):
    leads_per_day: int = 20
    industry_filters: list[str] = []


@router.put("/users/{user_id}/settings")
async def update_user_settings(
    user_id: int,
    body: UserSettingsRequest,
    current_user: dict = Depends(require_admin),
):
    """Update leads_per_day and industry_filters for a commercial user."""
    import json as _json
    async with get_conn() as conn:
        cursor = await conn.execute("SELECT id FROM lu_users WHERE id = ?", (user_id,))
        row = await cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")

        await conn.execute(
            "UPDATE lu_users SET leads_per_day = ?, industry_filters = ? WHERE id = ?",
            (body.leads_per_day, _json.dumps(body.industry_filters), user_id),
        )
        await conn.commit()
    return {"user_id": user_id, "leads_per_day": body.leads_per_day, "industry_filters": body.industry_filters}


@router.get("/analytics")
async def get_analytics(
    current_user: dict = Depends(require_admin),
):
    """Return analytics: total leads, by status, by commercial, conversion rate."""
    import json as _json
    today = str(date.today())

    async with get_conn() as conn:
        # Overall stats for today
        cursor = await conn.execute(
            "SELECT COUNT(*) FROM lu_daily_assignments WHERE assigned_date = ?", (today,)
        )
        row = await cursor.fetchone()
        total_today = row[0]

        # Breakdown by status (all time)
        cursor = await conn.execute(
            "SELECT status, COUNT(*) AS count FROM lu_daily_assignments GROUP BY status"
        )
        status_rows = await cursor.fetchall()

        # Breakdown by status today
        cursor = await conn.execute(
            "SELECT status, COUNT(*) AS count FROM lu_daily_assignments WHERE assigned_date = ? GROUP BY status",
            (today,),
        )
        status_today_rows = await cursor.fetchall()

        # Per commercial stats (including this week's closed leads)
        cursor = await conn.execute(
            """
            SELECT
                u.id,
                u.name,
                u.email,
                u.lead_search_enabled,
                COALESCE(u.leads_per_day, 20) AS leads_per_day,
                COALESCE(u.industry_filters, '[]') AS industry_filters,
                COUNT(da.id) AS total_assigned,
                COUNT(CASE WHEN da.status = 'closed' THEN 1 END) AS closed,
                COUNT(CASE WHEN da.status = 'no_answer' THEN 1 END) AS no_answer,
                COUNT(CASE WHEN da.status = 'rejected' THEN 1 END) AS rejected,
                COUNT(CASE WHEN da.status = 'pending' THEN 1 END) AS pending,
                COUNT(CASE WHEN da.status = 'call_later' THEN 1 END) AS call_later,
                COUNT(CASE WHEN da.status = 'wrong_number' THEN 1 END) AS wrong_number,
                COUNT(CASE WHEN da.assigned_date = ? THEN 1 END) AS today_count,
                COUNT(CASE WHEN da.status = 'closed' AND da.assigned_date::date >= CURRENT_DATE - INTERVAL '7 days' AND da.assigned_date::date < CURRENT_DATE THEN 1 END) AS week_closed
            FROM lu_users u
            LEFT JOIN lu_daily_assignments da ON da.user_id = u.id
            GROUP BY u.id, u.name, u.email, u.lead_search_enabled
            ORDER BY u.name
            """,
            (today,),
        )
        commercial_rows = await cursor.fetchall()

        # Top industries per commercial
        cursor = await conn.execute(
            """
            SELECT da.user_id, c.industry, COUNT(*) AS cnt
            FROM lu_daily_assignments da
            JOIN lu_companies c ON c.id = da.company_id
            WHERE c.industry IS NOT NULL AND c.industry != ''
            GROUP BY da.user_id, c.industry
            ORDER BY da.user_id, cnt DESC
            """
        )
        industry_rows = await cursor.fetchall()

        # Total companies in DB
        cursor = await conn.execute("SELECT COUNT(*) FROM lu_companies")
        row = await cursor.fetchone()
        total_companies = row[0]

    # Build top_industries dict per user
    top_industries: dict[int, list] = {}
    for r in industry_rows:
        uid = r["user_id"]
        if uid not in top_industries:
            top_industries[uid] = []
        if len(top_industries[uid]) < 5:
            top_industries[uid].append({"industry": r["industry"], "count": r["cnt"]})

    all_assigned = sum(r["total_assigned"] for r in commercial_rows)
    all_closed = sum(r["closed"] for r in commercial_rows)
    conversion_rate = round((all_closed / all_assigned * 100), 1) if all_assigned > 0 else 0

    status_map = {r["status"]: r["count"] for r in status_rows}
    status_today_map = {r["status"]: r["count"] for r in status_today_rows}

    # Get all users for assignment dropdown
    users_list = [
        {"id": r["id"], "name": r["name"], "email": r["email"]}
        for r in commercial_rows
    ]

    return {
        "today": str(today),
        "total_leads_today": total_today,
        "total_companies": total_companies,
        "users": users_list,
        "all_time": {
            "total_assigned": all_assigned,
            "by_status": status_map,
            "conversion_rate": conversion_rate,
        },
        "today_by_status": status_today_map,
        "by_commercial": [
            {
                "id": r["id"],
                "name": r["name"],
                "email": r["email"],
                "lead_search_enabled": bool(r["lead_search_enabled"]),
                "leads_per_day": r["leads_per_day"],
                "industry_filters": _json.loads(r["industry_filters"] or "[]"),
                "total_assigned": r["total_assigned"],
                "today_count": r["today_count"],
                "closed": r["closed"],
                "week_closed": r["week_closed"],
                "no_answer": r["no_answer"],
                "rejected": r["rejected"],
                "pending": r["pending"],
                "call_later": r["call_later"],
                "conversion_rate": round(
                    (r["closed"] / r["total_assigned"] * 100) if r["total_assigned"] > 0 else 0, 1
                ),
                "top_industries": top_industries.get(r["id"], []),
            }
            for r in commercial_rows
        ],
    }


@router.get("/commercial/{user_id}/leads")
async def get_commercial_leads(
    user_id: int,
    current_user: dict = Depends(require_admin),
):
    """Return all leads assigned to a commercial with full details."""
    async with get_conn() as conn:
        cursor = await conn.execute(
            """
            SELECT
                da.id, da.status, da.assigned_date,
                c.name AS company_name, c.city, c.industry,
                c.phone, c.phone2,
                da.call_notes, da.rejection_feedback
            FROM lu_daily_assignments da
            JOIN lu_companies c ON c.id = da.company_id
            WHERE da.user_id = ?
            ORDER BY da.assigned_date DESC, c.name
            LIMIT 200
            """,
            (user_id,),
        )
        rows = await cursor.fetchall()
        return {
            "leads": [
                {
                    "id": r["id"],
                    "status": r["status"],
                    "assigned_date": str(r["assigned_date"]),
                    "company_name": r["company_name"],
                    "city": r["city"] or "",
                    "industry": r["industry"] or "",
                    "phone": r["phone"] or "",
                    "phone2": r["phone2"] or "",
                    "call_notes": r["call_notes"] or "",
                    "rejection_feedback": r["rejection_feedback"] or "",
                }
                for r in rows
            ]
        }


@router.patch("/lead-search-toggle")
async def toggle_lead_search(
    body: LeadSearchToggle,
    current_user: dict = Depends(require_admin),
):
    """Enable or disable lead search for a specific user."""
    async with get_conn() as conn:
        cursor = await conn.execute(
            "SELECT id, name FROM lu_users WHERE id = ?", (body.user_id,)
        )
        row = await cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Usuario no encontrado")

        enabled_val = "TRUE" if body.enabled else "FALSE"
        await conn.execute(
            f"UPDATE lu_users SET lead_search_enabled = {enabled_val} WHERE id = ?",
            (body.user_id,),
        )
        await conn.commit()

    return {
        "user_id": body.user_id,
        "name": row["name"],
        "lead_search_enabled": body.enabled,
    }


@router.post("/trigger-enrichment")
async def trigger_enrichment(
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(require_admin),
):
    """Enrich companies that have no enrichment data yet."""

    async def _enrich_pending() -> None:
        import json

        async with get_conn() as conn:
            cursor = await conn.execute(
                """
                SELECT id, name, city, industry, website, phone
                FROM lu_companies
                WHERE opportunity_analysis IS NULL OR opportunity_analysis = ''
                LIMIT 50
                """
            )
            companies = await cursor.fetchall()

        for company in companies:
            try:
                enrichment = await enrich_company(dict(company))
                async with get_conn() as conn:
                    await conn.execute(
                        """
                        UPDATE lu_companies SET
                            digital_score = CASE WHEN COALESCE(digital_score, 0) > 0
                                            THEN digital_score ELSE ? END,
                            opportunity_level = CASE WHEN COALESCE(digital_score, 0) > 0
                                                THEN opportunity_level ELSE ? END,
                            redes_sociales = CASE WHEN COALESCE(digital_score, 0) > 0
                                             THEN redes_sociales ELSE ? END,
                            captacion_leads = CASE WHEN COALESCE(digital_score, 0) > 0
                                              THEN captacion_leads ELSE ? END,
                            email_marketing = CASE WHEN COALESCE(digital_score, 0) > 0
                                              THEN email_marketing ELSE ? END,
                            video_contenido = CASE WHEN COALESCE(digital_score, 0) > 0
                                              THEN video_contenido ELSE ? END,
                            seo_info = CASE WHEN COALESCE(digital_score, 0) > 0
                                       THEN seo_info ELSE ? END,
                            hooks = ?,
                            opening_lines = ?,
                            opportunity_analysis = ?
                        WHERE id = ?
                        """,
                        (
                            enrichment.get("digital_score", 30),
                            enrichment.get("opportunity_level", "media"),
                            1 if enrichment.get("redes_sociales", False) else 0,
                            1 if enrichment.get("captacion_leads", False) else 0,
                            1 if enrichment.get("email_marketing", False) else 0,
                            1 if enrichment.get("video_contenido", False) else 0,
                            1 if enrichment.get("seo_info", False) else 0,
                            json.dumps(enrichment.get("hooks", [])),
                            json.dumps(enrichment.get("opening_lines", [])),
                            enrichment.get("opportunity_analysis", ""),
                            company["id"],
                        ),
                    )
                    await conn.commit()
                logger.info(f"Enriched company id={company['id']}")
            except Exception as exc:
                logger.error(f"Enrichment failed for company {company['id']}: {exc}")

    background_tasks.add_task(_enrich_pending)
    return {"message": "Enriquecimiento iniciado en segundo plano"}


@router.post("/lusha-load")
async def lusha_load(
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(require_admin),
):
    """
    Fetch 25 leads from Lusha (construction/reform sector) and
    assign them round-robin among the active commercial users.
    """
    from services.lusha_leads import search_construction_leads
    from services.enrichment import enrich_company

    async def _load_task():
        import json as _json

        logger.info("Lusha load: starting 25-lead fetch")
        try:
            contacts = await search_construction_leads(count=25)
        except Exception as e:
            logger.error(f"Lusha search failed: {e}")
            return

        # Round-robin across active commercial users
        async with get_conn() as conn:
            rows = await (await conn.execute(
                "SELECT id FROM lu_users WHERE role = 'commercial' ORDER BY id"
            )).fetchall()
        target_users = [r["id"] for r in rows]
        if not target_users:
            logger.warning("Lusha load: no commercial users to assign to")
            return
        today = str(date.today())
        loaded = 0

        for i, c in enumerate(contacts):
            user_id = target_users[i % len(target_users)]
            try:
                # Create or find company
                async with get_conn() as conn:
                    # Dedup by name
                    cursor = await conn.execute(
                        "SELECT id FROM lu_companies WHERE name = ?",
                        (c["company_name"],),
                    )
                    existing = await cursor.fetchone()

                    if existing:
                        company_id = existing["id"]
                    else:
                        # Enrich with Claude
                        enrichment = await enrich_company({
                            "name": c["company_name"],
                            "city": c["company_city"],
                            "industry": c["company_industry"],
                            "website": c["company_website"],
                            "phone": "",
                        })

                        cursor = await conn.execute(
                            """
                            INSERT INTO lu_companies
                              (name, website, city, industry, digital_score, opportunity_level,
                               redes_sociales, captacion_leads, email_marketing,
                               video_contenido, seo_info, hooks, opening_lines, opportunity_analysis)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                            """,
                            (
                                c["company_name"],
                                c["company_website"],
                                c["company_city"],
                                c["company_industry"],
                                enrichment.get("digital_score", 30),
                                enrichment.get("opportunity_level", "media"),
                                1 if enrichment.get("redes_sociales") else 0,
                                1 if enrichment.get("captacion_leads") else 0,
                                1 if enrichment.get("email_marketing") else 0,
                                1 if enrichment.get("video_contenido") else 0,
                                1 if enrichment.get("seo_info") else 0,
                                _json.dumps(enrichment.get("hooks", [])),
                                _json.dumps(enrichment.get("opening_lines", [])),
                                enrichment.get("opportunity_analysis", ""),
                            ),
                        )
                        company_id = cursor.lastrowid
                        await conn.commit()

                        # Create contact (phone masked, not revealed)
                        if c.get("contact_name"):
                            await conn.execute(
                                """
                                INSERT INTO lu_contacts
                                  (company_id, name, title, email,
                                   lusha_person_id, phone_revealed, phone_prefix)
                                VALUES (?,?,?,?,?,0,?)
                                """,
                                (
                                    company_id,
                                    c["contact_name"],
                                    c.get("contact_title", ""),
                                    c.get("contact_email", ""),
                                    c.get("lusha_person_id", ""),
                                    c.get("phone_prefix", ""),
                                ),
                            )
                            await conn.commit()

                # Assign to user (ignore duplicate constraint)
                async with get_conn() as conn:
                    try:
                        await conn.execute(
                            """
                            INSERT INTO lu_daily_assignments
                              (company_id, user_id, assigned_date, status)
                            VALUES (?,?,?,?)
                            ON CONFLICT DO NOTHING
                            """,
                            (company_id, user_id, str(today), "pending"),
                        )
                        await conn.commit()
                        loaded += 1
                    except Exception as ex:
                        logger.warning(f"Duplicate assignment skipped: {ex}")

            except Exception as ex:
                logger.error(f"Error processing Lusha contact {i}: {ex}")
                continue

        logger.info(f"Lusha load complete: {loaded} leads assigned")

    background_tasks.add_task(_load_task)
    return {
        "message": "Lusha load started in background. 25 leads distributed across commercial users.",
    }


@router.post("/load-real-leads")
async def load_real_leads(
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(require_admin),
):
    """
    Scrape real leads from Google Maps via Apify, enrich with Claude,
    delete demo/test leads and assign real ones round-robin to commercial users.
    """
    from services.apify_gmaps import fetch_real_leads
    from services.enrichment import enrich_company

    async def _task():
        import json as _json
        from datetime import date as _date

        logger.info("Starting real lead load via Apify + Claude enrichment")

        # 1. Fetch real leads from Google Maps
        try:
            leads = await fetch_real_leads(per_search=8)
        except Exception as e:
            logger.error(f"Apify fetch failed: {e}")
            return

        if not leads:
            logger.error("No leads returned from Apify")
            return

        logger.info(f"Apify returned {len(leads)} real leads")

        # 2. Wipe all existing demo/test data (only non-real companies)
        async with get_conn() as conn:
            # Delete demo assignments (companies with no website are demo/test markers)
            await conn.execute(
                "DELETE FROM lu_daily_assignments WHERE company_id IN "
                "(SELECT id FROM lu_companies WHERE website IS NULL OR website = '' OR website LIKE 'www.%' OR rating IS NULL)"
            )
            # Delete contacts and companies that have no website (demo/test data markers)
            await conn.execute(
                "DELETE FROM lu_contacts WHERE company_id IN "
                "(SELECT id FROM lu_companies WHERE website IS NULL OR website = '' OR website LIKE 'www.%' OR rating IS NULL)"
            )
            await conn.execute(
                "DELETE FROM lu_companies WHERE website IS NULL OR website = '' OR website LIKE 'www.%' OR rating IS NULL"
            )
            await conn.commit()
            logger.info("Wiped fake/test leads from database")

        # 3. Insert real leads + enrich + assign
        async with get_conn() as conn:
            rows = await (await conn.execute(
                "SELECT id FROM lu_users WHERE role = 'commercial' ORDER BY id"
            )).fetchall()
        target_users = [r["id"] for r in rows]
        if not target_users:
            logger.warning("load-real-leads: no commercial users to assign to")
            return
        today = str(_date.today())
        loaded = 0

        for i, lead in enumerate(leads[:30]):
            user_id = target_users[i % len(target_users)]
            try:
                # Dedup by phone
                async with get_conn() as conn:
                    cur = await conn.execute(
                        "SELECT id FROM lu_companies WHERE phone = ?", (lead["phone"],)
                    )
                    existing = await cur.fetchone()
                    if existing:
                        company_id = existing["id"]
                    else:
                        # Enrich with Claude
                        try:
                            enrichment = await enrich_company({
                                "name": lead["name"],
                                "phone": lead["phone"],
                                "website": lead.get("website", ""),
                                "city": lead["city"],
                                "industry": lead["industry"],
                            })
                        except Exception as ex:
                            logger.warning(f"Enrichment failed for {lead['name']}: {ex}")
                            enrichment = {}

                        cur = await conn.execute(
                            """INSERT INTO lu_companies
                              (name, phone, website, city, industry, rating, reviews_count,
                               digital_score, opportunity_level,
                               redes_sociales, captacion_leads, email_marketing,
                               video_contenido, seo_info, hooks, opening_lines, opportunity_analysis)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                            (
                                lead["name"],
                                lead["phone"],
                                lead.get("website", ""),
                                lead["city"],
                                lead["industry"],
                                lead.get("rating"),
                                lead.get("reviews_count"),
                                enrichment.get("digital_score", 40),
                                enrichment.get("opportunity_level", "media"),
                                1 if enrichment.get("redes_sociales") else 0,
                                1 if enrichment.get("captacion_leads") else 0,
                                1 if enrichment.get("email_marketing") else 0,
                                1 if enrichment.get("video_contenido") else 0,
                                1 if enrichment.get("seo_info") else 0,
                                _json.dumps(enrichment.get("hooks", [])),
                                _json.dumps(enrichment.get("opening_lines", [])),
                                enrichment.get("opportunity_analysis", ""),
                            ),
                        )
                        company_id = cur.lastrowid
                        await conn.commit()

                        # Contact with phone from Google Maps (public info)
                        await conn.execute(
                            """INSERT INTO lu_contacts (company_id, name, phone, title, phone_revealed)
                            VALUES (?, ?, ?, ?, 1)""",
                            (company_id, lead["name"], lead["phone"], "Responsable"),
                        )
                        await conn.commit()

                # Assign to user
                async with get_conn() as conn:
                    await conn.execute(
                        """INSERT INTO lu_daily_assignments
                           (company_id, user_id, assigned_date, status)
                           VALUES (?,?,?,?)
                           ON CONFLICT DO NOTHING""",
                        (company_id, user_id, today, "pending"),
                    )
                    await conn.commit()
                loaded += 1

            except Exception as ex:
                logger.error(f"Error loading lead {lead.get('name')}: {ex}")

        logger.info(f"Real lead load complete: {loaded}/{len(leads)} leads inserted")

    background_tasks.add_task(_task)
    return {
        "message": "Real lead load started (Google Maps + Claude). Takes ~5 minutes.",
        "note": "Demo/test leads will be removed and replaced with real leads.",
    }


@router.post("/scrape-now", tags=["admin"])
async def scrape_now(
    background_tasks: BackgroundTasks,
    current_user: dict = Depends(require_admin),
):
    """Trigger Google Maps scraping immediately (admin only)."""

    from services.google_maps_leads import scrape_and_enrich_leads

    async def _scrape_task():
        logger.info("Google Maps scraping initiated by admin")
        try:
            stats = await scrape_and_enrich_leads()
            logger.info(f"Scraping complete: {stats}")
        except Exception as e:
            logger.error(f"Scraping error: {e}")

    background_tasks.add_task(_scrape_task)
    return {"message": "Scraping iniciado en segundo plano"}


@router.post("/seed-leads")
async def seed_leads(
    current_user: dict = Depends(require_admin),
):
    """Load 24 demo leads (12 constructoras + 12 abogados) for today."""
    import json as _json

    constructoras = [
        ("Construcciones Aranda S.L.", "Madrid", "Construcción", "www.construccionesaranda.es", "+34 915 234 567", 25, "alta"),
        ("Reformas Blanca & Cia", "Barcelona", "Reformas Integrales", "www.reformasblanca.com", "+34 932 456 789", 18, "alta"),
        ("Constructor Díaz Valencia", "Valencia", "Obras Civiles", "constructordiaz.es", "+34 963 567 890", 32, "media"),
        ("Obras Mayores Sevilla", "Sevilla", "Construcción de Viviendas", "www.obrmayor.es", "+34 954 678 901", 21, "alta"),
        ("Constructora Gómez & Partners", "Bilbao", "Construcción Comercial", "gomezbuilders.es", "+34 944 789 012", 28, "media"),
        ("Reformas García Hermanos", "Málaga", "Reformas y Rehabilitación", "www.reformasgarcia.com", "+34 952 890 123", 15, "alta"),
        ("Obras Públicas Castro", "Zaragoza", "Obras de Infraestructura", "www.obrascastro.es", "+34 976 901 234", 26, "media"),
        ("Constructora Jiménez Toledo", "Toledo", "Edificación", "jimenezconstructora.es", "+34 925 012 345", 19, "alta"),
        ("Reformas Inteligentes Alicante", "Alicante", "Reformas Domésticas", "www.reformasinteligentes.es", "+34 965 123 456", 22, "media"),
        ("Constructor Mora Granada", "Granada", "Construcción Residencial", "moraconstructor.com", "+34 958 234 567", 17, "alta"),
        ("Obras Urbanas Murcia", "Murcia", "Urbanización", "www.obrasurbanas.es", "+34 968 345 678", 24, "media"),
        ("Constructora Sánchez Santander", "Santander", "Construcción General", "sanchez-construcciones.es", "+34 942 456 789", 20, "alta"),
    ]

    abogados = [
        ("Bufete Jurídico García López", "Madrid", "Derecho Mercantil", "www.garcialopezabogados.es", "+34 913 456 789", 35, "media"),
        ("Abogados Martínez & Asociados", "Barcelona", "Derecho Penal", "www.martinez-abogados.cat", "+34 932 567 890", 28, "alta"),
        ("Despacho Fernández Valencia", "Valencia", "Derecho Laboral", "fernandezabogados.es", "+34 963 678 901", 31, "media"),
        ("Bufete Sánchez Sevilla", "Sevilla", "Derecho Civil", "www.bufetesanchez.es", "+34 954 789 012", 22, "alta"),
        ("Abogados Ramírez & Cia", "Bilbao", "Derecho Mercantil", "ramirezabogados.es", "+34 944 890 123", 33, "media"),
        ("Despacho Jurídico López Málaga", "Málaga", "Derecho Inmobiliario", "www.lopezabogados.com", "+34 952 901 234", 25, "alta"),
        ("Bufete Colón Zaragoza", "Zaragoza", "Derecho Administrativo", "www.bufetecolon.es", "+34 976 012 345", 29, "media"),
        ("Abogados Ruiz Toledo", "Toledo", "Derecho de Familia", "ruizabogados.es", "+34 925 123 456", 24, "alta"),
        ("Despacho Gómez Alicante", "Alicante", "Derecho Mercantil", "www.gomezabogados.es", "+34 965 234 567", 32, "media"),
        ("Bufete Granada Legal", "Granada", "Derecho Penal", "www.granadalegal.com", "+34 958 345 678", 26, "alta"),
        ("Abogados Morales Murcia", "Murcia", "Derecho Civil", "moralesabogados.es", "+34 968 456 789", 30, "media"),
        ("Despacho Jiménez Santander", "Santander", "Derecho Laboral", "www.jimenezabogados.es", "+34 942 567 890", 27, "alta"),
    ]

    all_clients = constructoras + abogados
    today = str(date.today())
    loaded = 0

    async with get_conn() as conn:
        await conn.execute("DELETE FROM lu_daily_assignments WHERE assigned_date = ?", (today,))
        await conn.commit()

        rows = await (await conn.execute(
            "SELECT id FROM lu_users WHERE role = 'commercial' ORDER BY id"
        )).fetchall()
        user_list = [r["id"] for r in rows]
        if not user_list:
            return {"message": "No commercial users to assign demo leads to", "loaded": 0}

        for i, (name, city, industry, website, phone, score, opp) in enumerate(all_clients):
            hooks = _json.dumps([
                f"Sector {industry} en {city}",
                f"Score digital {score}/100",
                f"Oportunidad {opp}"
            ])
            opening = _json.dumps([
                f"Hola {name}, soy [YOUR NAME] de [YOUR COMPANY]",
                f"Buenos días, le llamo porque hemos analizado {name}",
                f"{name}, ¿tiene 3 minutos?"
            ])

            cursor = await conn.execute(
                """INSERT INTO lu_companies
                (name, website, city, industry, phone, digital_score, opportunity_level, hooks, opening_lines)
                VALUES (?,?,?,?,?,?,?,?,?)
                ON CONFLICT DO NOTHING""",
                (name, website, city, industry, phone, score, opp, hooks, opening)
            )
            await conn.commit()

            cursor = await conn.execute("SELECT id FROM lu_companies WHERE name = ?", (name,))
            row = await cursor.fetchone()
            company_id = row["id"]

            contact_names = ["Carlos Ruiz", "María Fernández", "Juan García", "Patricia López", "Roberto Sánchez", "Elena Torres"]
            contact = contact_names[i % len(contact_names)]

            await conn.execute("DELETE FROM lu_contacts WHERE company_id = ?", (company_id,))
            await conn.execute(
                """INSERT INTO lu_contacts (company_id, name, title, email, phone_prefix, lusha_person_id, phone_revealed)
                VALUES (?,?,?,?,?,?,?)""",
                (company_id, contact, "Director", f"{contact.lower().replace(' ', '.')}@{name.lower().replace(' ', '')}.es", str(630 + i % 10), f"mock_{i:03d}", 0)
            )
            await conn.commit()

            user_id = user_list[i % len(user_list)]
            await conn.execute(
                """INSERT INTO lu_daily_assignments (company_id, user_id, assigned_date, status)
                VALUES (?,?,?,?)
                ON CONFLICT DO NOTHING""",
                (company_id, user_id, today, "pending")
            )
            await conn.commit()
            loaded += 1

    return {
        "message": f"✅ {loaded} demo leads loaded",
        "date": today,
        "assigned_to_users": len(user_list),
        "sectors": ["12 construction firms", "12 law firms"]
    }


@router.get("/unassigned-leads")
async def get_unassigned_leads(current_user: dict = Depends(require_admin)):
    """Get all companies without assignments, grouped by industry (niche)."""
    async with get_conn() as conn:
        cursor = await conn.execute("""
            SELECT c.id, c.name, c.city, c.industry, c.phone, COUNT(lda.id) as assigned_count
            FROM lu_companies c
            LEFT JOIN lu_daily_assignments lda ON c.id = lda.company_id
            WHERE lda.id IS NULL
            GROUP BY c.id, c.industry
            ORDER BY c.industry, c.name
        """)
        rows = await cursor.fetchall()

        niche_groups = {}
        for row in rows:
            industry = row['industry'] or 'Sin clasificar'
            if industry not in niche_groups:
                niche_groups[industry] = []
            niche_groups[industry].append({
                'id': row['id'],
                'name': row['name'],
                'city': row['city'],
                'phone': row['phone'],
            })

        summary = {niche: len(leads) for niche, leads in niche_groups.items()}
        total = sum(len(leads) for leads in niche_groups.values())

        return {
            'total_unassigned': total,
            'by_niche': niche_groups,
            'summary': summary,
        }


class BulkAssignRequest(BaseModel):
    assignments: list[dict]  # [{'company_id': int, 'user_id': int}, ...]


@router.post("/assign-bulk")
async def assign_bulk(
    body: BulkAssignRequest,
    current_user: dict = Depends(require_admin),
):
    """Bulk assign unassigned companies to users."""
    async with get_conn() as conn:
        assigned_count = 0
        errors = []

        for assignment in body.assignments:
            company_id = assignment.get('company_id')
            user_id = assignment.get('user_id')

            if not company_id or not user_id:
                errors.append(f"Invalid assignment: {assignment}")
                continue

            try:
                await conn.execute("""
                    INSERT INTO lu_daily_assignments
                    (company_id, user_id, status)
                    VALUES (?, ?, 'pending')
                    ON CONFLICT DO NOTHING
                """, (company_id, user_id))
                await conn.commit()
                assigned_count += 1
            except Exception as e:
                errors.append(f"Company {company_id}: {str(e)}")

        return {
            'success': True,
            'assigned': assigned_count,
            'errors': errors if errors else None,
        }


@router.get("/rejection-analysis")
async def get_rejection_analysis(
    industry: Optional[str] = None,
    user_id: Optional[int] = None,
    current_user: dict = Depends(get_current_user),
):
    """Analyze rejection feedback with Claude to produce failure points. Accessible to all users."""
    import json as _json
    import anthropic
    from config import get_settings

    cache_key = f"{industry or '*'}:{user_id or '*'}"
    now = time.time()
    if cache_key in _rejection_cache:
        ts, cached = _rejection_cache[cache_key]
        if now - ts < _REJECTION_CACHE_TTL:
            return cached

    async with get_conn() as conn:
        sql = """
            SELECT
                da.rejection_feedback,
                c.industry,
                u.name AS commercial_name
            FROM lu_daily_assignments da
            JOIN lu_companies c ON c.id = da.company_id
            JOIN lu_users u ON u.id = da.user_id
            WHERE da.status = 'rejected'
              AND da.rejection_feedback IS NOT NULL
              AND TRIM(da.rejection_feedback) != ''
        """
        params: list = []
        if industry:
            sql += " AND c.industry = ?"
            params.append(industry)
        if user_id:
            sql += " AND da.user_id = ?"
            params.append(user_id)
        sql += " ORDER BY da.assigned_date DESC LIMIT 100"

        cursor = await conn.execute(sql, tuple(params))
        rows = await cursor.fetchall()

    if not rows:
        result = {
            "failure_points": [],
            "summary": "Sin datos suficientes aún. Los puntos de fracaso aparecerán conforme los comerciales registren rechazos.",
            "total_feedback": 0,
        }
        _rejection_cache[cache_key] = (now, result)
        return result

    feedbacks = [r["rejection_feedback"] for r in rows]
    context_parts = []
    if industry:
        context_parts.append(f"Sector: {industry}")
    if user_id and rows:
        context_parts.append(f"Comercial: {rows[0]['commercial_name']}")
    context_label = " | ".join(context_parts) if context_parts else "Todos los sectores y comerciales"

    feedback_block = "\n".join(f"- {t}" for t in feedbacks[:60])

    prompt = f"""Eres un analista de ventas B2B especializado en identificar patrones de fracaso comercial.

Contexto: {context_label}
Total de registros de rechazo analizados: {len(feedbacks)}

Feedbacks registrados por los comerciales cuando una empresa fue rechazada:
{feedback_block}

Analiza estos feedbacks e identifica los principales PUNTOS DE FRACASO en el proceso comercial.
Responde ÚNICAMENTE con un JSON válido con esta estructura exacta (sin texto adicional):
{{
  "failure_points": [
    {{"title": "Nombre corto (3-5 palabras)", "description": "Explicación clara en 1-2 frases de qué está fallando y por qué", "frequency": "alta|media|baja", "count": número_aproximado_de_casos}},
    ...
  ],
  "summary": "Resumen ejecutivo en 2-3 frases de los patrones principales y qué deberían mejorar los comerciales"
}}

Identifica entre 3 y 6 puntos de fracaso distintos. Sé específico, actionable y basado solo en los datos."""

    settings = get_settings()
    client = anthropic.Anthropic(api_key=settings.anthropic_api_key)

    try:
        _msgs = [{"role": "user", "content": prompt}]
        message = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=1024,
            messages=_msgs,
        )
        track_anthropic_call(
            name="leadup.admin.rejection_analysis",
            model="claude-haiku-4-5-20251001",
            messages=_msgs,
            response=message,
        )
        content = message.content[0].text
        start = content.find('{')
        end = content.rfind('}') + 1
        if start >= 0 and end > start:
            data = _json.loads(content[start:end])
        else:
            data = {"failure_points": [], "summary": content}
    except Exception as exc:
        logger.error(f"Rejection analysis error: {exc}")
        data = {"failure_points": [], "summary": "Error al analizar el feedback. Inténtalo de nuevo."}

    result = {**data, "total_feedback": len(feedbacks)}
    _rejection_cache[cache_key] = (now, result)
    return result


@router.get("/pending-by-niche")
async def get_pending_by_niche(
    current_user: dict = Depends(require_admin),
):
    """Empresas sin asignar a nadie, agrupadas por nicho."""
    async with get_conn() as conn:
        rows = await (await conn.execute(
            """
            SELECT
                COALESCE(NULLIF(TRIM(c.industry), ''), 'Sin nicho') AS industry,
                COUNT(*) AS available
            FROM lu_companies c
            WHERE c.id NOT IN (SELECT company_id FROM lu_daily_assignments)
            GROUP BY COALESCE(NULLIF(TRIM(c.industry), ''), 'Sin nicho')
            ORDER BY available DESC
            """
        )).fetchall()

    niches = [
        {"industry": r["industry"], "total_pending": r["available"]}
        for r in rows
        if r["available"] > 0
    ]

    return {"niches": niches}


# ── Exportar notas a Excel ─────────────────────────────────────────────────────

@router.get("/export-notes")
async def export_notes(
    current_user: dict = Depends(require_admin),
):
    """Export all assignments with company info, contact, notes and reminders as Excel."""
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    STATUS_LABELS = {
        "pending":    "Pendiente",
        "no_answer":  "Sin respuesta",
        "call_later": "Llamar luego",
        "closed":     "Agendado",
        "rejected":   "Rechazado",
    }

    async with get_conn() as conn:
        cursor = await conn.execute(
            """
            SELECT
                da.id              AS assignment_id,
                da.assigned_date,
                da.status,
                da.notes,
                da.follow_up_date,
                da.rejection_feedback,
                u.name             AS comercial,
                c.name             AS empresa,
                c.city,
                c.industry         AS sector,
                c.website,
                c.phone            AS empresa_tel,
                c.digital_score,
                c.opportunity_level,
                ct.name            AS contacto_nombre,
                ct.title           AS contacto_cargo,
                ct.phone           AS contacto_tel,
                ct.email           AS contacto_email
            FROM lu_daily_assignments da
            JOIN lu_users u    ON u.id  = da.user_id
            JOIN lu_companies c ON c.id = da.company_id
            LEFT JOIN LATERAL (
                SELECT name, title, phone, email
                FROM lu_contacts
                WHERE company_id = c.id
                ORDER BY id
                LIMIT 1
            ) ct ON true
            ORDER BY da.assigned_date DESC, u.name, c.name
            """
        )
        rows = await cursor.fetchall()

        reminder_cursor = await conn.execute(
            """
            SELECT assignment_id, text, due_at, done
            FROM lu_reminders
            ORDER BY assignment_id, position ASC, created_at ASC
            """
        )
        reminder_rows = await reminder_cursor.fetchall()

    # Group reminders by assignment_id
    reminders_map: dict = {}
    for r in reminder_rows:
        aid = r["assignment_id"]
        if aid not in reminders_map:
            reminders_map[aid] = []
        tick = "✓" if r["done"] else "○"
        due = f" [{r['due_at']}]" if r["due_at"] else ""
        reminders_map[aid].append(f"{tick} {r['text']}{due}")

    # Build workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Notas LeadUp"

    HEADERS = [
        "Fecha Asignación", "Comercial", "Estado",
        "Empresa", "Ciudad", "Sector", "Web", "Tel. Empresa",
        "Score Digital", "Oportunidad",
        "Contacto", "Cargo", "Tel. Contacto", "Email",
        "Notas", "Seguimiento", "Recordatorios", "Feedback Rechazo",
    ]

    # Header style
    header_fill   = PatternFill("solid", fgColor="3B1F6E")   # brand violet
    header_font   = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    center_align  = Alignment(horizontal="center", vertical="center", wrap_text=False)
    wrap_align    = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    thin          = Side(border_style="thin", color="D0D0D0")
    cell_border   = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[1].height = 22
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align
        cell.border    = cell_border

    # Column widths
    COL_WIDTHS = [14, 14, 14, 30, 14, 22, 28, 16, 8, 12, 22, 18, 16, 26, 40, 14, 40, 30]
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    STATUS_COLORS = {
        "closed":     "D1FAE5",
        "pending":    "FEF3C7",
        "call_later": "EDE9FE",
        "no_answer":  "F1F5F9",
        "rejected":   "FEE2E2",
    }

    for row_idx, r in enumerate(rows, start=2):
        aid    = r["assignment_id"]
        status = r["status"] or "pending"
        fill   = PatternFill("solid", fgColor=STATUS_COLORS.get(status, "FFFFFF"))
        reminders_text = "\n".join(reminders_map.get(aid, [])) or ""

        values = [
            str(r["assigned_date"]) if r["assigned_date"] else "",
            r["comercial"] or "",
            STATUS_LABELS.get(status, status),
            r["empresa"] or "",
            r["city"] or "",
            r["sector"] or "",
            r["website"] or "",
            r["empresa_tel"] or "",
            r["digital_score"] if r["digital_score"] is not None else "",
            r["opportunity_level"] or "",
            r["contacto_nombre"] or "",
            r["contacto_cargo"] or "",
            r["contacto_tel"] or "",
            r["contacto_email"] or "",
            r["notes"] or "",
            str(r["follow_up_date"]) if r["follow_up_date"] else "",
            reminders_text,
            r["rejection_feedback"] or "",
        ]

        has_content = bool(r["notes"] or reminders_text or r["rejection_feedback"])
        row_height  = 32 if has_content else 18
        ws.row_dimensions[row_idx].height = row_height

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = fill
            cell.border    = cell_border
            cell.alignment = wrap_align if col_idx in (15, 17, 18) else Alignment(vertical="center")

    # Freeze header row
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from datetime import datetime as _dt
    filename = f"leadup_notas_{_dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
